#!/usr/bin/env python3
"""Enrich the raw FAA metrics CSV into a manager-facing "proof" workbook + docs.

The goal of this step is to produce evidence that the Failure Analysis Agent
(FAA) captures *real* issues. It selects the substantive finds (see
``is_substantive``) and, for each, combs the agent's full ``report.md`` with an
LLM to write a short, factual, manager-readable narrative:

    what failed  ->  what FAA found  ->  why that was valuable

It writes:

  * ``faa-metrics-organized.xlsx`` - three sheets:
      - "Real Issues"  : curated, manager-facing proof rows (AI narrative +
                         a derived value tag + a blank managerNotes column).
      - "All Runs"     : the full raw dataset, for traceability.
      - "Summary"      : aggregate proof metrics.
  * ``faa-metrics-real-issues.md`` - the curated highlights as a paste-ready
    Markdown doc (drops straight into Word / Confluence / a wiki).
  * ``faa-metrics-summary.md``     - the aggregate metrics as Markdown.

AI enrichment is optional. If the ``AZURE_OPENAI_*`` env vars are absent, or
``--enable-ai false`` is passed, the narrative columns are left blank and the
step still succeeds - matching the FAA's own degrade-gracefully behavior.

Env (inherited from the pipeline's linked variable group):
  AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_DEPLOYMENT,
  AZURE_OPENAI_API_KEY, AZURE_OPENAI_API_VERSION
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import common

# The full report.md can be large (embedded evidence/log excerpts); cap what we
# send to the model so a single incident stays within a sane token budget.
REPORT_CHAR_BUDGET = 16000

AI_SYSTEM_PROMPT = (
    "You are reviewing the output of an automated CI failure-analysis agent (FAA) "
    "for Azure Container Networking, to help an engineering manager see the concrete "
    "value it delivered. You are given (a) structured fields for one analyzed failure "
    "and (b) the agent's full markdown report. Write a factual, non-marketing "
    "assessment grounded ONLY in the report - never invent details. Respond ONLY with "
    "a JSON object with keys 'whatHappened', 'faaFinding', and 'faaValue'. "
    "'whatHappened' = the failure/scenario in one plain sentence a manager understands. "
    "'faaFinding' = the agent's root-cause conclusion, concrete and specific (name the "
    "component, error, or test). 'faaValue' = why this catch is a real, useful result: "
    "e.g. it identified a genuine product regression, pinpointed a known flake and its "
    "fix, correctly reassigned ownership, or overturned an incorrect initial guess. "
    "Each value <= 500 characters. Use empty strings only if the report truly lacks the "
    "information."
)

# Structured fields handed to the model alongside the report body.
AI_FIELDS = [
    "category", "confidence", "confidenceBand", "analysisStatus", "classificationSource",
    "rootCauseSummary", "recommendedOwner", "recommendedAction", "proposedFix",
    "verdict", "falsificationOutcome", "nodeAssessment", "clusterName", "os", "cni",
]

AI_NARRATIVE_COLUMNS = ["whatHappened", "faaFinding", "faaValue"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP_COLUMNS = {
    "rootCauseSummary", "recommendedAction", "proposedFix", "verdict", "nodeAssessment",
    "whatHappened", "faaFinding", "faaValue", "valueTag", "managerNotes",
}


def log(msg: str) -> None:
    print(f"[enrich] {msg}", flush=True)


# --------------------------------------------------------------------------- #
# Curation: which rows are "real issues", and what makes them valuable.
# --------------------------------------------------------------------------- #
def _norm(row: dict, key: str) -> str:
    return (row.get(key, "") or "").strip()


def is_substantive(row: dict) -> bool:
    """A "real issue" worth showing a manager.

    Requires that the agent actually produced a classification, that it landed a
    root-cause conclusion, and that it carries at least one impact signal: a
    concrete proposed fix, a high-confidence verdict, or an overturned pre-match
    ('refuted' falsification). Flakes count - a flake the agent correctly
    identified is a real, useful result.
    """
    status = _norm(row, "analysisStatus").lower()
    if status == "analysis_failed":
        return False
    has_root = bool(_norm(row, "rootCauseSummary") or _norm(row, "verdict"))
    if not has_root:
        return False
    has_signal = (
        bool(_norm(row, "proposedFix"))
        or _norm(row, "confidenceBand").lower() == "high"
        or _norm(row, "falsificationOutcome").lower() == "refuted"
    )
    return has_signal


def value_tags(row: dict) -> list[str]:
    """Deterministic, evidence-based tags describing why a find is valuable."""
    tags: list[str] = []
    if _norm(row, "falsificationOutcome").lower() == "refuted":
        tags.append("Overturned initial guess")
    category = _norm(row, "category")
    label = common.CATEGORY_LABELS.get(category)
    if label:
        tags.append(label)
    if _norm(row, "proposedFix"):
        tags.append("Proposed fix")
    if _norm(row, "confidenceBand").lower() == "high":
        tags.append("High confidence")
    # De-dupe while preserving order.
    seen: set[str] = set()
    return [t for t in tags if not (t in seen or seen.add(t))]


def value_score(row: dict) -> int:
    """Rank curated rows so the most compelling proof floats to the top."""
    score = 0
    if _norm(row, "falsificationOutcome").lower() == "refuted":
        score += 4
    if _norm(row, "confidenceBand").lower() == "high":
        score += 2
    if _norm(row, "proposedFix"):
        score += 2
    if _norm(row, "category") == "pr_regression":
        score += 1
    return score


# --------------------------------------------------------------------------- #
# AI enrichment (reads the full report.md).
# --------------------------------------------------------------------------- #
class AzureOpenAI:
    def __init__(self, endpoint, deployment, api_key, api_version):
        self.url = (
            f"{endpoint.rstrip('/')}/openai/deployments/{deployment}"
            f"/chat/completions?api-version={api_version}"
        )
        self.headers = {"Content-Type": "application/json", "api-key": api_key}

    def summarize(self, row: dict, report_text: str) -> dict:
        facts = {k: row.get(k, "") for k in AI_FIELDS}
        excerpt = report_text[:REPORT_CHAR_BUDGET]
        user_content = (
            "Structured fields:\n"
            + json.dumps(facts, ensure_ascii=False)
            + "\n\n----- FULL REPORT (markdown) -----\n"
            + excerpt
        )
        body = {
            "messages": [
                {"role": "system", "content": AI_SYSTEM_PROMPT},
                {"role": "user", "content": user_content},
            ],
            "response_format": {"type": "json_object"},
        }
        resp = requests.post(self.url, headers=self.headers, json=body, timeout=120)
        if resp.status_code >= 400:
            detail = resp.text[:300].replace("\n", " ")
            raise ValueError(f"HTTP {resp.status_code}: {detail}")
        content = resp.json()["choices"][0]["message"]["content"]
        parsed = json.loads(content)
        return {
            "whatHappened": str(parsed.get("whatHappened", "")).strip(),
            "faaFinding": str(parsed.get("faaFinding", "")).strip(),
            "faaValue": str(parsed.get("faaValue", "")).strip(),
        }


def build_ai_client(enable_ai: bool):
    if not enable_ai:
        log("AI enrichment disabled via --enable-ai false; narrative columns left blank.")
        return None
    endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT", "").strip()
    deployment = os.environ.get("AZURE_OPENAI_DEPLOYMENT", "").strip()
    api_key = os.environ.get("AZURE_OPENAI_API_KEY", "").strip()
    api_version = os.environ.get("AZURE_OPENAI_API_VERSION", "").strip()
    if not (endpoint and deployment and api_key and api_version):
        log("AZURE_OPENAI_* not fully configured; narrative columns left blank.")
        return None
    log(f"AI enrichment enabled (deployment '{deployment}').")
    return AzureOpenAI(endpoint, deployment, api_key, api_version)


def read_report(raw_dir: str, row: dict) -> str:
    folder = _norm(row, "sourceFolder")
    if not raw_dir or not folder:
        return ""
    path = os.path.join(raw_dir, folder, common.REPORT_FILENAME)
    if not os.path.isfile(path):
        return ""
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        return fh.read()


def enrich_curated(rows: list[dict], client, raw_dir: str) -> None:
    """Populate value tags (always) and AI narrative (when configured)."""
    for row in rows:
        row["valueTag"] = "; ".join(value_tags(row))
        row["managerNotes"] = ""
        for col in AI_NARRATIVE_COLUMNS:
            row.setdefault(col, "")
        if client is None:
            continue
        report_text = read_report(raw_dir, row)
        try:
            row.update(client.summarize(row, report_text))
        except (requests.RequestException, KeyError, ValueError) as exc:
            log(f"build {row.get('buildId', '?')}: AI enrichment failed: {exc}")


# --------------------------------------------------------------------------- #
# Summary metrics (framed as proof of captured value).
# --------------------------------------------------------------------------- #
def compute_summary(rows, curated, from_date, to_date) -> list[tuple[str, object]]:
    total = len(rows)
    status = Counter(_norm(r, "analysisStatus") or "unknown" for r in rows)
    analyzed = status.get("analyzed", 0)

    cat = Counter(_norm(r, "category") or "unknown" for r in curated)
    refuted = sum(1 for r in curated if _norm(r, "falsificationOutcome").lower() == "refuted")
    with_fix = sum(1 for r in curated if _norm(r, "proposedFix"))
    high_conf = sum(1 for r in curated if _norm(r, "confidenceBand").lower() == "high")

    summary: list[tuple[str, object]] = [
        ("Date range", f"{from_date} to {to_date}"),
        ("Failures FAA processed", total),
        ("Successfully analyzed (produced a classification)", analyzed),
        ("Real issues captured (substantive finds)", len(curated)),
        ("  - with a concrete proposed fix", with_fix),
        ("  - high confidence", high_conf),
        ("  - AI overturned the initial guess (refuted)", refuted),
        ("", ""),
        ("Real issues by category", ""),
    ]
    for name, count in cat.most_common():
        label = common.CATEGORY_LABELS.get(name, name)
        summary.append((f"  {label}", count))
    return summary


# --------------------------------------------------------------------------- #
# Output writers.
# --------------------------------------------------------------------------- #
def style_header(ws, ncols) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _write_sheet(ws, rows, columns) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    style_header(ws, len(columns))
    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        if col in WRAP_COLUMNS:
            ws.column_dimensions[letter].width = 50
            for cell in ws[letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            ws.column_dimensions[letter].width = max(12, min(len(col) + 4, 28))


def write_workbook(path, curated, all_rows, summary) -> None:
    wb = Workbook()
    real = wb.active
    real.title = "Real Issues"
    _write_sheet(real, curated, common.CURATED_COLUMNS)

    all_ws = wb.create_sheet("All Runs")
    _write_sheet(all_ws, all_rows, common.RAW_COLUMNS)

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Metric", "Value"])
    for name, value in summary:
        summary_ws.append([name, value])
    style_header(summary_ws, 2)
    summary_ws.column_dimensions["A"].width = 52
    summary_ws.column_dimensions["B"].width = 28

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


def _md_escape(text: str) -> str:
    return (text or "").replace("|", "\\|").replace("\n", " ").strip()


def write_real_issues_md(path, curated, from_date, to_date) -> None:
    """Paste-ready highlights doc for a manager."""
    captured = len(curated)
    with_fix = sum(1 for r in curated if _norm(r, "proposedFix"))
    refuted = sum(1 for r in curated if _norm(r, "falsificationOutcome").lower() == "refuted")

    lines = [
        "# FAA - Real Issues Captured",
        "",
        f"_Failure Analysis Agent findings, {from_date} to {to_date}._",
        "",
        (
            f"Over this window the agent captured **{captured}** substantive issues - "
            f"**{with_fix}** with a concrete proposed fix and **{refuted}** where it "
            f"overturned the deterministic pre-match. Each row links to the full report."
        ),
        "",
        "| Date | Scenario | Category | Value | What happened | What FAA found | Why it mattered | Report |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for r in curated:
        scenario = _md_escape(
            " / ".join(x for x in (_norm(r, "pipelineName"), _norm(r, "clusterName")) if x)
            or _norm(r, "stage")
        )
        category = _md_escape(common.CATEGORY_LABELS.get(_norm(r, "category"), _norm(r, "category")))
        url = _norm(r, "reportUrl")
        link = f"[link]({url})" if url else ""
        lines.append(
            "| {date} | {scenario} | {category} | {value} | {what} | {finding} | {why} | {link} |".format(
                date=_md_escape(_norm(r, "buildDate")[:10]),
                scenario=scenario,
                category=category,
                value=_md_escape(_norm(r, "valueTag")),
                what=_md_escape(_norm(r, "whatHappened") or _norm(r, "rootCauseSummary")),
                finding=_md_escape(_norm(r, "faaFinding") or _norm(r, "rootCauseSummary")),
                why=_md_escape(_norm(r, "faaValue")),
                link=link,
            )
        )
    if not curated:
        lines.append("| _no substantive findings in range_ | | | | | | | |")
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")


def write_summary_md(path, summary) -> None:
    lines = ["# FAA Success Metrics", "", "| Metric | Value |", "| --- | --- |"]
    for name, value in summary:
        if not name and value == "":
            continue
        lines.append(f"| {name.rstrip() or '&nbsp;'} | {value} |")
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--in-csv", required=True)
    parser.add_argument("--raw-dir", default="", help="raw/ tree from the collect step")
    parser.add_argument("--out-xlsx", required=True)
    parser.add_argument("--out-md", required=True, help="aggregate metrics markdown")
    parser.add_argument("--out-issues-md", required=True, help="paste-ready highlights doc")
    parser.add_argument("--from-date", default="")
    parser.add_argument("--to-date", default="")
    parser.add_argument("--enable-ai", default="true")
    args = parser.parse_args()

    enable_ai = str(args.enable_ai).strip().lower() in ("true", "1", "yes")
    rows = common.read_csv_rows(args.in_csv) if os.path.isfile(args.in_csv) else []
    log(f"loaded {len(rows)} row(s) from {args.in_csv}")

    curated = [r for r in rows if is_substantive(r)]
    curated.sort(key=lambda r: (value_score(r), _norm(r, "buildDate")), reverse=True)
    log(f"{len(curated)} substantive 'real issue' row(s) selected")

    client = build_ai_client(enable_ai)
    enrich_curated(curated, client, args.raw_dir)

    summary = compute_summary(rows, curated, args.from_date, args.to_date)

    write_workbook(args.out_xlsx, curated, rows, summary)
    write_real_issues_md(args.out_issues_md, curated, args.from_date, args.to_date)
    write_summary_md(args.out_md, summary)
    log(f"wrote {args.out_xlsx}, {args.out_issues_md}, and {args.out_md}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
